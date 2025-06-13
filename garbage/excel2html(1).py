import re
from openpyxl import load_workbook
from openpyxl.cell import MergedCell


class to_html():
    def __init__(self, file, save_file, sheet_name=None):
        self.file = file  # 文件路径
        self.save_file = save_file  # html保存路径
        self.sheet_name = sheet_name  # sheet名

    def creat_html(self):
        wb = load_workbook(filename=self.file)
        # 如果没有传sheet_name，默认取第一个sheet
        if self.sheet_name is None:
            sheet = wb[wb.sheetnames[0]]
        else:
            sheet = wb[self.sheet_name]

        # === 新增：找出第一个和最后一个有数据的行和列 ===
        first_data_row = None
        last_data_row = 0
        first_data_col = None
        last_data_col = 0

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if first_data_row is None or cell.row < first_data_row:
                        first_data_row = cell.row
                    if cell.row > last_data_row:
                        last_data_row = cell.row
                    if first_data_col is None or cell.column < first_data_col:
                        first_data_col = cell.column
                    if cell.column > last_data_col:
                        last_data_col = cell.column

        if first_data_row is None:
            # 没有任何数据
            with open(self.save_file, 'w', encoding='utf-8') as f:
                f.write('<table border="1"><tr><td>无数据</td></tr></table>')
            return

        # === 原有逻辑保持不变，仅调整遍历范围为有效区域 ===
        cell_dic = {}  # 用于储存 所有合并单元格的左上单元格对象
        col_width = {}  # 用于储存 所有列的列宽,px
        row_height = {}  # 用于储存 所有列的行高,px

        # 查询列宽
        for col in sheet.columns:
            pat = re.compile(r"[A-Z]+")
            colname = pat.findall(col[0].coordinate)[0]
            px = round(sheet.column_dimensions[colname].width * 5)
            col_width[colname] = px

        # 查询行高
        for row in sheet.rows:
            pat = re.compile(r"[A-Z]+(\d+)")
            rowid = int(pat.findall(row[0].coordinate)[0])
            px = sheet.row_dimensions[rowid].height
            if px is None:
                px = 13.5
            row_height[str(rowid)] = px

        # 处理合并单元格
        for merged_range in sheet.merged_cells.ranges:
            now_width = 0
            now_height = 0
            for i in range(merged_range.min_col, merged_range.max_col + 1):
                coord = sheet.cell(row=1, column=i).coordinate
                colname = re.compile(r"[A-Z]+").findall(coord)[0]
                now_width += col_width.get(colname, 0)

            for i in range(merged_range.min_row, merged_range.max_row + 1):
                coord = sheet.cell(row=i, column=1).coordinate
                colindex = re.compile(r"[A-Z]+(\d+)").findall(coord)[0]
                now_height += row_height.get(colindex, 0)

            now_width = int(now_width)
            now_height = int(now_height)

            cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            colspan = merged_range.max_col - merged_range.min_col + 1
            rowspan = merged_range.max_row - merged_range.min_row + 1
            cell_dic[cell] = (now_height, now_width, colspan, rowspan)

        html = '''<table border="1">'''

        # 裁剪到有效数据区域
        rows = sheet.iter_rows(
            min_row=first_data_row,
            max_row=last_data_row,
            min_col=first_data_col,
            max_col=last_data_col
        )

        for row in rows:
            tr = '''<tr>'''
            for cell in row:
                td = ""
                if cell in cell_dic:
                    text = cell.value or ''
                    vertical = f'vertical-align: {cell.alignment.vertical};' if cell.alignment.vertical else ''
                    horizontal = f'text-align: {cell.alignment.horizontal};' if cell.alignment.horizontal else ''

                    font_size = str(int(cell.font.size) + 3)
                    font_weight = '700' if cell.font.b else '400'
                    style = f'"color: rgb(0, 0, 0); font-size: {font_size}px; font-weight: {font_weight}; font-style: normal;{vertical}{horizontal}"'

                    height, width, colspan, rowspan = cell_dic[cell]
                    td = f'''<td height="{height}" width="{width}" colspan="{colspan}" rowspan="{rowspan}" style={style}>{text}</td>'''
                elif not isinstance(cell, MergedCell):
                    vertical = f'vertical-align: {cell.alignment.vertical};' if cell.alignment.vertical else ''
                    horizontal = f'text-align: {cell.alignment.horizontal};' if cell.alignment.horizontal else ''

                    coord = cell.coordinate
                    match = re.compile(r"([A-Z]+)(\d+)").match(coord)
                    if match:
                        colname, rowid = match.groups()
                        height = row_height.get(rowid, '')
                        width = col_width.get(colname, '')

                        font_size = str(int(cell.font.size) + 3)
                        font_weight = '700' if cell.font.b else '400'
                        style = f'"color: rgb(0, 0, 0); font-size: {font_size}px; font-weight: {font_weight}; font-style: normal;{vertical}{horizontal}"'

                        td = f'<td height="{height}" width="{width}" style={style}>{cell.value or ""}</td>'
                tr += td
            tr += '</tr>'
            html += tr

        html += '</table>'
        with open(self.save_file, 'w', encoding='utf-8') as f:
            f.write(html)


if __name__ == '__main__':
    data = to_html(r"D:\asianInfo\ExcelAssist\（7.12）2024年自然灾害对巩固拓展脱贫攻坚成果工作的影响情况表.xlsx",
                    r"D:\asianInfo\ExcelAssist\（7.12）2024年自然灾害对巩固拓展脱贫攻坚成果工作的影响情况表.html")
    data.creat_html()
